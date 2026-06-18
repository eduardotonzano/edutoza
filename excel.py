"""Relatorio Excel: NOTICIAS + CARTEIRA (todos os 340 ativos c/ status) + EMISSORES."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime

_HF   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_HFILL= PatternFill("solid", start_color="1F3864")
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT  = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="CCCCCC")
_BD   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FA   = PatternFill("solid", start_color="EBF1F8")   # zebra A
_FB   = PatternFill("solid", start_color="FFFFFF")    # zebra B
_FVERDE = PatternFill("solid", start_color="D6EFD6")  # noticia encontrada
_FVERM  = PatternFill("solid", start_color="FFE0E0")  # sem noticia hoje
_FCINZA = PatternFill("solid", start_color="ECECEC")  # sem noticia aplicavel

_STATUS_FILL = {"Noticia encontrada": _FVERDE, "Sem noticia hoje": _FVERM,
                "Sem noticia aplicavel": _FCINZA}


def _cab(ws, cols, larg):
    for c, t in enumerate(cols, 1):
        x = ws.cell(row=1, column=c, value=t)
        x.font=_HF; x.fill=_HFILL; x.alignment=_CTR; x.border=_BD
    for c, l in enumerate(larg, 1):
        ws.column_dimensions[get_column_letter(c)].width = l
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def gerar_excel(alvos, noticias_finais, caminho=None):
    """alvos: lista dos 340 ativos (de mapeamento.mapear_carteira), cada um com
    ativo/tipo/categoria/emissor/news_applicable. noticias_finais: itens de todas as fontes."""
    contagem = Counter(n["nome"] for n in noticias_finais)   # noticias por EMISSOR (nome)
    melhor_fonte = {}
    for n in noticias_finais:
        melhor_fonte.setdefault(n["nome"], n["fonte"])

    wb = Workbook()

    # ---- Aba 1: NOTICIAS ----
    ws1 = wb.active; ws1.title = "NOTICIAS"
    _cab(ws1, ["DATA","TICKER","EMISSOR","TITULO","FONTE","LINK","RELEVANCIA","RESUMO IA","IMPACTO"],
         [18,14,24,60,20,45,11,55,16])
    ta=None; grp=True
    for i, n in enumerate(noticias_finais, 2):
        if n["ticker"] != ta: grp = not grp; ta = n["ticker"]
        fill = _FA if grp else _FB
        for c, v in enumerate([n["data"],n["ticker"],n["nome"],n["titulo"],n["fonte"],
                               n["link"],n["relevancia"],n["resumo_ia"],n["impacto"]], 1):
            x = ws1.cell(row=i, column=c, value=v)
            x.font=Font(name="Arial",size=10); x.fill=fill; x.border=_BD; x.alignment=_LFT
        if n["link"]:
            ws1.cell(row=i, column=6).hyperlink = n["link"]
            ws1.cell(row=i, column=6).font = Font(name="Arial", size=10, color="0563C1", underline="single")

    # ---- Aba 2: CARTEIRA (todos os 340 ativos com status) ----
    ws2 = wb.create_sheet("CARTEIRA")
    _cab(ws2, ["ATIVO","TIPO","CATEGORIA","EMISSOR","STATUS","N NOTICIAS","MELHOR FONTE"],
         [42,18,12,26,22,11,22])
    for i, a in enumerate(alvos, 2):
        q = contagem.get(a["emissor"], 0)
        if not a["news_applicable"]:
            status = "Sem noticia aplicavel"
        elif q > 0:
            status = "Noticia encontrada"
        else:
            status = "Sem noticia hoje"
        fill = _STATUS_FILL[status]
        for c, v in enumerate([a["ativo"], a["tipo"], a["categoria"], a["emissor"],
                               status, q, melhor_fonte.get(a["emissor"], "-")], 1):
            x = ws2.cell(row=i, column=c, value=v)
            x.font=Font(name="Arial",size=10); x.fill=fill; x.border=_BD; x.alignment=_LFT

    # ---- Aba 3: EMISSORES (um por emissor monitorado, com contagem) ----
    ws3 = wb.create_sheet("EMISSORES")
    _cab(ws3, ["EMISSOR","N ATIVOS","N NOTICIAS","MELHOR FONTE"], [28,11,12,22])
    monit = {}
    for a in alvos:
        if a["news_applicable"] and a["emissor"] != "-":
            monit[a["emissor"]] = monit.get(a["emissor"], 0) + 1
    linhas = sorted(monit.items(), key=lambda kv: (-contagem.get(kv[0], 0), kv[0]))
    for i, (emissor, n_ativos) in enumerate(linhas, 2):
        q = contagem.get(emissor, 0)
        fill = _FVERDE if q > 0 else (_FA if i % 2 == 0 else _FB)
        for c, v in enumerate([emissor, n_ativos, q, melhor_fonte.get(emissor, "-")], 1):
            x = ws3.cell(row=i, column=c, value=v)
            x.font=Font(name="Arial",size=10); x.fill=fill; x.border=_BD; x.alignment=_LFT

    if caminho is None:
        caminho = f"Noticias_Carteira_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
    wb.save(caminho)
    return caminho, len(noticias_finais), len([a for a in alvos if a["news_applicable"]])
